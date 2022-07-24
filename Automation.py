from pydoc import text
from PIL import ImageGrab
from matplotlib.pyplot import title
import pywinauto,time
from pywinauto.application import Application
from pywinauto.keyboard import send_keys
from pywinauto import mouse
from pytesseract import pytesseract
import pyautogui as pag
from PIL import Image
import sys
from openpyxl import workbook, load_workbook



app=Application().start(r"C:\Users\FHZ\AppData\Local\Programs\ITDe-Filing_6\ITDe-Filing_6.exe")
time.sleep(4)

pag.press(['enter', 'tab', 'enter'])
time.sleep(3)

pywinauto.mouse.double_click(button='left', coords=(400, 550))
time.sleep(3)

pywinauto.mouse.click(button='left', coords=(1500,550))
time.sleep(3)

pywinauto.mouse.click(button='left', coords=(550,750))
time.sleep(3)

pywinauto.mouse.click(button='left', coords=(1086,928))
time.sleep(3)

pag.moveTo(349, 547, 2)
pag.click()
time.sleep(2)

pywinauto.mouse.click(button='left', coords=(317,169))
time.sleep(2)
pag.press('enter')
time.sleep(7)

ss_region = (270, 576, 962, 635)
ss_img = ImageGrab.grab(ss_region)
ss_img.save("S1.jpg")
path_to_tesseract = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
image_path1 = r"C:\Users\FHZ\Downloads\Auto\s1.jpg"

img1 = Image.open(image_path1)
pytesseract.tesseract_cmd = path_to_tesseract
text1 = pytesseract.image_to_string(img1)
print(text1[:-1])

wb=load_workbook('Result.xlsx')
ws=wb.active
for cell in ws["C"]:
    if cell.value is None:
        cell.value = text1
    
wb.save('Result.xlsx')
string1=text1
string2="Error: Error in importing, Please upload proper online downloaded draft JSON file."
if string1 == string2 :
    sys.exit()


pywinauto.mouse.click(button='left', coords=(1100,721))
time.sleep(2)

pywinauto.mouse.click(button='left', coords=(1574,907))
time.sleep(3)
pag.press(['enter', 'tab', 'enter'])
time.sleep(1)

pag.press('end')
time.sleep(1)
pywinauto.mouse.click(button='left', coords=(402,326))
time.sleep(3)

pag.press('end')
time.sleep(1)
pywinauto.mouse.click(button='left', coords=(1544,812))
time.sleep(2)

pag.press('end')
time.sleep(1)
pywinauto.mouse.click(button='left', coords=(1608,653))
time.sleep(3)

pywinauto.mouse.click(button='left', coords=(313,667))
time.sleep(2)
pag.typewrite('Gurgaon')
time.sleep(2)

pywinauto.mouse.click(button='left', coords=(285,810))
time.sleep(1)
pag.press('end')
pywinauto.mouse.click(button='left', coords=(1514,767))
time.sleep(3)
pag.press('end')
time.sleep(1)

ss2_region = (243,168,1667,572)
ss2_img = ImageGrab.grab(ss2_region)
ss2_img.save("S2.jpg")


image_path2 = r"C:\Users\FHZ\Downloads\Auto\s2.jpg"
img2 = Image.open(image_path2)
pytesseract.tesseract_cmd = path_to_tesseract
text2 = pytesseract.image_to_string(img2)
print(text2[:-1])


wb=load_workbook('Result.xlsx')
ws=wb.active
for cell in ws["D"]:
    if cell.value is None:
        cell.value = text2
wb.save('Result.xlsx')
